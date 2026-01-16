"""
API REST para Sistema InfoTaxi con Swagger UI
Servicios de gestión de usuarios y reportes
"""

from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
from flasgger import Swagger, swag_from
import mysql.connector
from mysql.connector import Error
import hashlib
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os
import secrets
from functools import wraps

app = Flask(__name__)

# Configurar CORS de manera más permisiva
# Importante: Incluir todos los headers que Swagger UI necesita
# Configuración global de CORS para todas las rutas
CORS(app, 
     resources={
         r"/*": {
             "origins": "*",
             "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
             "allow_headers": ["Content-Type", "X-User-Celular", "Authorization", "Accept"],
             "supports_credentials": False,
             "max_age": 3600
         }
     },
     supports_credentials=False)

# ==================== CONFIGURACIÓN SWAGGER ====================
swagger_config = {
    "headers": [],
    "specs": [
        {
            "endpoint": 'apispec',
            "route": '/apispec.json',
            "rule_filter": lambda rule: True,
            "model_filter": lambda tag: True,
        }
    ],
    "static_url_path": "/flasgger_static",
    "swagger_ui": True,
    "specs_route": "/apidocs/",
    "openapi": "2.0"
}

# Configuración dinámica de Swagger
# El host se detectará automáticamente desde la petición
swagger_template = {
    "swagger": "2.0",
    "info": {
        "title": "API InfoTaxi",
        "description": "Sistema de gestión de reportes de conductores de taxi",
        "version": "1.0.0",
        "contact": {
            "name": "Soporte API",
            "email": "jandrezapata@hotmail.com"
        }
    },
    # No especificar host para que use el host actual de la petición
    "basePath": "/",
    "schemes": ["http", "https"],
    "securityDefinitions": {
        "CelularAuth": {
            "type": "apiKey",
            "name": "X-User-Celular",
            "in": "header",
            "description": "Número de celular del usuario autenticado"
        }
    },
    "consumes": ["application/json"],
    "produces": ["application/json"]
}

# Inicializar Swagger con configuración dinámica
swagger = Swagger(app, config=swagger_config, template=swagger_template)

# CRÍTICO: Usar after_request para interceptar y modificar la respuesta de /apispec.json
# Esto asegura que siempre agreguemos el campo 'host' incluso si Flasgger no lo incluye
@app.after_request
def add_host_to_swagger_spec(response):
    """Interceptar respuesta de /apispec.json y agregar el campo 'host' si falta"""
    if request.path == '/apispec.json' and response.status_code == 200:
        try:
            # Obtener el JSON de la respuesta
            import json
            data = response.get_json()
            
            if data and isinstance(data, dict):
                # Detectar host y scheme
                current_host = request.host
                current_scheme = request.scheme
                
                # Si hay un header X-Forwarded-Host (proxy), usarlo
                forwarded_host = request.headers.get('X-Forwarded-Host')
                if forwarded_host:
                    current_host = forwarded_host.split(',')[0].strip()
                
                # Si hay un header X-Forwarded-Proto (proxy), usarlo
                forwarded_proto = request.headers.get('X-Forwarded-Proto')
                if forwarded_proto:
                    current_scheme = forwarded_proto.split(',')[0].strip()
                
                # Si el host contiene el dominio de Easypanel, asegurar que no tenga puerto
                if 'easypanel.host' in current_host or 'fxtfoe.easypanel.host' in current_host:
                    if ':' in current_host:
                        host_parts = current_host.split(':')
                        current_host = host_parts[0]
                    if current_scheme == 'http' and 'easypanel' in current_host:
                        current_scheme = 'https'
                
                # FORZAR el host siempre - esto es crítico para Swagger UI
                data['host'] = current_host
                data['schemes'] = [current_scheme]
                
                print(f"[SWAGGER] Host agregado: {data['host']}, Scheme: {data['schemes']}")
                
                # Actualizar la respuesta
                response.data = json.dumps(data).encode('utf-8')
                response.headers['Content-Type'] = 'application/json'
        except Exception as e:
            print(f"[SWAGGER ERROR] Error agregando host: {str(e)}")
            import traceback
            traceback.print_exc()
    
    return response

# También registrar el endpoint directamente por si acaso
@app.route('/apispec.json', methods=['GET', 'OPTIONS'], endpoint='custom_swagger_spec')
def get_swagger_spec():
    """Obtener especificación de Swagger con host dinámico basado en la petición"""
    print(f"[SWAGGER] Endpoint personalizado llamado - Método: {request.method}")
    
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,Accept,X-User-Celular,Authorization")
        response.headers.add('Access-Control-Allow-Methods', "GET,OPTIONS")
        response.headers.add('Access-Control-Max-Age', "3600")
        return response
    
    try:
        # Obtener la especificación base de Swagger
        spec = swagger.get_apispecs()
        print(f"[SWAGGER] Especificación obtenida, tiene host: {'host' in spec}")
        
        # Actualizar con el host y scheme actuales de la petición
        current_host = request.host
        current_scheme = request.scheme
        
        # Si hay un header X-Forwarded-Host (proxy), usarlo
        forwarded_host = request.headers.get('X-Forwarded-Host')
        if forwarded_host:
            current_host = forwarded_host.split(',')[0].strip()
        
        # Si hay un header X-Forwarded-Proto (proxy), usarlo
        forwarded_proto = request.headers.get('X-Forwarded-Proto')
        if forwarded_proto:
            current_scheme = forwarded_proto.split(',')[0].strip()
        
        # Si el host contiene el dominio de Easypanel, asegurar que no tenga puerto
        if 'easypanel.host' in current_host or 'fxtfoe.easypanel.host' in current_host:
            if ':' in current_host:
                host_parts = current_host.split(':')
                current_host = host_parts[0]
            if current_scheme == 'http' and 'easypanel' in current_host:
                current_scheme = 'https'
        
        # FORZAR el host siempre - esto es crítico
        spec['host'] = current_host
        spec['schemes'] = [current_scheme]
        
        print(f"[SWAGGER] Host final: {spec['host']}, Scheme: {spec['schemes']}")
        print(f"[SWAGGER] Verificación - host en spec: {'host' in spec}, valor: {spec.get('host')}")
        
        response = jsonify(spec)
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Accept,X-User-Celular,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,OPTIONS')
        response.headers.add('Access-Control-Max-Age', '3600')
        
        return response
    except Exception as e:
        print(f"[SWAGGER ERROR] {str(e)}")
        import traceback
        traceback.print_exc()
        error_response = jsonify({
            'error': str(e),
            'message': 'Error generando especificación Swagger',
            'host': request.host
        })
        error_response.headers.add('Access-Control-Allow-Origin', '*')
        return error_response, 500

# Manejar preflight requests (OPTIONS)
@app.before_request
def handle_preflight():
    if request.method == "OPTIONS":
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization,Accept")
        response.headers.add('Access-Control-Allow-Methods', "GET,PUT,POST,DELETE,OPTIONS")
        response.headers.add('Access-Control-Max-Age', "3600")
        return response

# Manejar preflight requests
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization,Accept')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    response.headers.add('Access-Control-Max-Age', '3600')
    return response

# Manejar errores globales (solo para excepciones no capturadas)
@app.errorhandler(500)
def handle_500_error(e):
    """Manejar errores 500"""
    print(f"[ERROR 500] Error del servidor: {str(e)}")
    import traceback
    traceback.print_exc()
    response = jsonify({
        'success': False,
        'message': f'Error del servidor: {str(e)}'
    })
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response, 500

@app.errorhandler(400)
def handle_bad_request(e):
    """Manejar errores de solicitud incorrecta"""
    return jsonify({
        'success': False,
        'message': 'Solicitud incorrecta: ' + str(e)
    }), 400

# ==================== CONFIGURACIÓN BASE DE DATOS ====================
DB_CONFIG = {
    'host': '31.97.130.20',
    'port': 4646,
    'database': 'electo',
    'user': 'mariadb',
    'password': '9204a8246f7ed4fe49e6'
}

def get_db_connection():
    """Crear conexión a la base de datos"""
    try:
        connection = mysql.connector.connect(**DB_CONFIG)
        return connection
    except Error as e:
        print(f"Error conectando a MySQL: {e}")
        return None

# ==================== DECORADOR DE AUTENTICACIÓN ====================
def verificar_usuario(f):
    """Decorador para verificar que el usuario existe por número de celular"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        celular = request.headers.get('X-User-Celular')
        
        if not celular:
            return jsonify({
                'success': False,
                'message': 'Número de celular requerido en headers (X-User-Celular)'
            }), 401
        
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Error de conexión a BD'}), 500
        
        try:
            cursor = conn.cursor(dictionary=True)
            cursor.execute(
                "SELECT id_user, username, nombres, rol FROM users WHERE Celular = %s AND isactive = 1",
                (celular,)
            )
            usuario = cursor.fetchone()
            
            if not usuario:
                return jsonify({
                    'success': False,
                    'message': 'Usuario no encontrado o inactivo'
                }), 403
            
            request.usuario = usuario
            
        except Error as e:
            return jsonify({'success': False, 'message': str(e)}), 500
        finally:
            if conn.is_connected():
                cursor.close()
                conn.close()
        
        return f(*args, **kwargs)
    
    return decorated_function

# ==================== SERVICIO 1: VERIFICAR USUARIO ====================
@app.route('/api/verificar-usuario', methods=['POST'])
def verificar_usuario_existe():
    """
    Verificar si un usuario existe por número de celular
    ---
    tags:
      - Usuarios
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          required:
            - celular
          properties:
            celular:
              type: string
              example: "3007471199"
              description: Número de celular del usuario
    responses:
      200:
        description: Usuario encontrado o no encontrado
        schema:
          type: object
          properties:
            success:
              type: boolean
            exists:
              type: boolean
            usuario:
              type: object
              properties:
                id:
                  type: integer
                nombre:
                  type: string
                email:
                  type: string
                rol:
                  type: string
                activo:
                  type: boolean
      400:
        description: Parámetros incorrectos
      500:
        description: Error del servidor
    """
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Error al procesar JSON: ' + str(e)
        }), 400
    
    if not data:
        return jsonify({
            'success': False,
            'message': 'No se recibió ningún dato JSON'
        }), 400
    
    celular = data.get('celular')
    
    if not celular:
        return jsonify({
            'success': False,
            'message': 'Número de celular requerido'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT id_user, username, nombres, Celular, rol, isactive 
            FROM users 
            WHERE Celular = %s
        """, (celular,))
        
        usuario = cursor.fetchone()
        
        if usuario:
            return jsonify({
                'success': True,
                'exists': True,
                'usuario': {
                    'id': usuario['id_user'],
                    'nombre': usuario['nombres'],
                    'email': usuario['username'],
                    'rol': usuario['rol'],
                    'activo': bool(usuario['isactive'])
                }
            }), 200
        else:
            return jsonify({
                'success': True,
                'exists': False,
                'message': 'Usuario no encontrado'
            }), 200
            
    except Error as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 2: CREAR USUARIO ====================
@app.route('/api/usuarios', methods=['POST', 'OPTIONS'])
def crear_usuario():
    """
    Crear nuevo usuario con rol 'usuario'
    ---
    tags:
      - Usuarios
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          required:
            - username
            - nombres
            - celular
            - password
          properties:
            username:
              type: string
              example: "usuario@ejemplo.com"
              description: Email del usuario
            nombres:
              type: string
              example: "Juan Pérez"
              description: Nombre completo
            celular:
              type: string
              example: "3001234567"
              description: Número de celular
            password:
              type: string
              example: "contraseña123"
              description: Contraseña del usuario
    responses:
      201:
        description: Usuario creado exitosamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
            id_user:
              type: integer
      400:
        description: Parámetros incorrectos
      409:
        description: Usuario ya existe
      500:
        description: Error del servidor
    """
    try:
        print(f"[DEBUG] POST /api/usuarios - Método: {request.method}")
        print(f"[DEBUG] Headers: {dict(request.headers)}")
        print(f"[DEBUG] Content-Type: {request.content_type}")
        
        if request.method == 'OPTIONS':
            response = jsonify({'status': 'ok'})
            response.headers.add("Access-Control-Allow-Origin", "*")
            response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization")
            response.headers.add('Access-Control-Allow-Methods', "GET,PUT,POST,DELETE,OPTIONS")
            return response
        
        try:
            print(f"[DEBUG] Intentando obtener JSON...")
            data = request.get_json(force=True)
            print(f"[DEBUG] JSON recibido: {data}")
        except Exception as e:
            print(f"[ERROR] Error al procesar JSON: {str(e)}")
            import traceback
            traceback.print_exc()
            response = jsonify({
                'success': False,
                'message': 'Error al procesar JSON: ' + str(e)
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 400
        
        if not data:
            print("[ERROR] No se recibió ningún dato JSON")
            response = jsonify({
                'success': False,
                'message': 'No se recibió ningún dato JSON'
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 400
        
        print(f"[DEBUG] Validando campos requeridos...")
        required = ['username', 'nombres', 'celular', 'password']
        if not all(field in data for field in required):
            missing = [field for field in required if field not in data]
            print(f"[ERROR] Campos faltantes: {missing}")
            response = jsonify({
                'success': False,
                'message': f'Campos requeridos faltantes: {", ".join(missing)}'
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 400
        
        print(f"[DEBUG] Conectando a la base de datos...")
        conn = get_db_connection()
        if not conn:
            print("[ERROR] No se pudo conectar a la base de datos")
            response = jsonify({'success': False, 'message': 'Error de conexión a la base de datos'})
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 500
        
        print(f"[DEBUG] Conexión a BD exitosa")
        
        try:
            print(f"[DEBUG] Creando cursor...")
            cursor = conn.cursor()
            
            # Verificar si el usuario ya existe
            print(f"[DEBUG] Verificando si el usuario ya existe...")
            cursor.execute("SELECT id_user FROM users WHERE Celular = %s OR username = %s",
                          (data['celular'], data['username']))
            existing = cursor.fetchone()
            if existing:
                print(f"[WARN] Usuario ya existe: {existing}")
                response = jsonify({
                    'success': False,
                    'message': 'Ya existe un usuario con ese celular o email'
                })
                response.headers.add('Access-Control-Allow-Origin', '*')
                response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization,Accept')
                response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
                return response, 409
            
            # Validar que los campos no estén vacíos
            if not data['username'] or not data['nombres'] or not data['celular'] or not data['password']:
                print(f"[ERROR] Campos vacíos detectados")
                response = jsonify({
                    'success': False,
                    'message': 'Todos los campos son requeridos y no pueden estar vacíos'
                })
                response.headers.add('Access-Control-Allow-Origin', '*')
                return response, 400
            
            # Hash de la contraseña
            print(f"[DEBUG] Generando hash de contraseña...")
            password_hash = hashlib.sha1(data['password'].encode()).hexdigest()
            
            # Insertar nuevo usuario
            print(f"[DEBUG] Insertando nuevo usuario...")
            query = """
                INSERT INTO users (username, nombres, Celular, rol, password, isactive, ultima_cone, ip, token)
                VALUES (%s, %s, %s, 'usuario', %s, 1, NOW(), '0.0.0.0', '')
            """
            cursor.execute(query, (
                data['username'],
                data['nombres'],
                data['celular'],
                password_hash
            ))
            conn.commit()
            
            user_id = cursor.lastrowid
            print(f"[SUCCESS] Usuario creado exitosamente con ID: {user_id}")
            
            response = jsonify({
                'success': True,
                'message': 'Usuario creado exitosamente',
                'id_user': user_id
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization,Accept')
            response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
            response.headers.add('Access-Control-Max-Age', '3600')
            return response, 201
            
        except Error as e:
            print(f"[ERROR] Error de base de datos: {str(e)}")
            import traceback
            traceback.print_exc()
            if conn and conn.is_connected():
                conn.rollback()
            response = jsonify({
                'success': False,
                'message': f'Error de base de datos: {str(e)}'
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 500
        except Exception as e:
            print(f"[ERROR] Error inesperado: {str(e)}")
            import traceback
            traceback.print_exc()
            if conn and conn.is_connected():
                conn.rollback()
            response = jsonify({
                'success': False,
                'message': f'Error inesperado: {str(e)}'
            })
            response.headers.add('Access-Control-Allow-Origin', '*')
            return response, 500
        finally:
            if conn and conn.is_connected():
                cursor.close()
                conn.close()
                print(f"[DEBUG] Conexión cerrada")
                
    except Exception as e:
        # Capturar cualquier error no manejado
        print(f"[ERROR CRÍTICO] Error no capturado en crear_usuario: {str(e)}")
        import traceback
        traceback.print_exc()
        response = jsonify({
            'success': False,
            'message': f'Error crítico: {str(e)}'
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response, 500

# ==================== SERVICIO 3A: CONSULTAR MIS REPORTES POR CÉDULA ====================
@app.route('/api/mis-reportes/<cedula>', methods=['GET'])
@verificar_usuario
def consultar_mis_reportes(cedula):
    """
    Consultar reportes propios por número de cédula
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    parameters:
      - name: cedula
        in: path
        type: string
        required: true
        description: Número de documento de identidad
        example: "8497643"
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
    responses:
      200:
        description: Resultados de la búsqueda
      401:
        description: No autorizado
      500:
        description: Error del servidor
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Buscar solo reportes creados por este usuario
        cursor.execute("""
            SELECT id, Fecha_Reporte, Numero_Documento, Nombres, Apellidos,
                   Fecha_cierre, Placa, Valor_Reporte, Descripcion_Reporte,
                   Vehiculo_afiliado, Estado
            FROM personas
            WHERE Numero_Documento = %s AND Reportante_Nombres = %s
            ORDER BY Fecha_Reporte DESC
        """, (cedula, request.usuario['id_user']))
        
        reportes = cursor.fetchall()
        
        if not reportes:
            return jsonify({
                'success': True,
                'found': False,
                'message': 'No tienes reportes creados para esta cédula',
                'reportes': []
            }), 200
        
        # Formatear respuesta
        reportes_formateados = []
        for r in reportes:
            reportes_formateados.append({
                'id': r['id'],
                'fecha_reporte': r['Fecha_Reporte'].strftime('%Y-%m-%d %H:%M') if r['Fecha_Reporte'] else None,
                'numero_documento': r['Numero_Documento'],
                'nombres': r['Nombres'],
                'apellidos': r['Apellidos'],
                'placa': r['Placa'],
                'valor_reporte': r['Valor_Reporte'],
                'descripcion': r['Descripcion_Reporte'],
                'estado': r['Estado']
            })
        
        return jsonify({
            'success': True,
            'found': True,
            'total_reportes': len(reportes),
            'reportes': reportes_formateados
        }), 200
        
    except Error as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 3: CONSULTAR PERSONA POR CÉDULA ====================
@app.route('/api/personas/<cedula>', methods=['GET'])
@verificar_usuario
def consultar_persona(cedula):
    """
    Consultar persona reportada por número de cédula
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    parameters:
      - name: cedula
        in: path
        type: string
        required: true
        description: Número de documento de identidad
        example: "8497643"
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
    responses:
      200:
        description: Resultados de la búsqueda
        schema:
          type: object
          properties:
            success:
              type: boolean
            found:
              type: boolean
            total_reportes:
              type: integer
            total_consultas:
              type: integer
              description: Total de consultas realizadas por el usuario
            reportes:
              type: array
              items:
                type: object
                properties:
                  id:
                    type: integer
                  fecha_reporte:
                    type: string
                  numero_documento:
                    type: string
                  nombres:
                    type: string
                  apellidos:
                    type: string
                  placa:
                    type: string
                  valor_reporte:
                    type: integer
                  descripcion:
                    type: string
                  estado:
                    type: string
      401:
        description: No autorizado
      403:
        description: Usuario inactivo
      500:
        description: Error del servidor
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Buscar persona
        cursor.execute("""
            SELECT id, Fecha_Reporte, Numero_Documento, Nombres, Apellidos,
                   Fecha_cierre, Placa, Valor_Reporte, Descripcion_Reporte,
                   Vehiculo_afiliado, Estado, Reportante_Nombres
            FROM personas
            WHERE Numero_Documento = %s
            ORDER BY Fecha_Reporte DESC
        """, (cedula,))
        
        reportes = cursor.fetchall()
        
        if not reportes:
            return jsonify({
                'success': True,
                'found': False,
                'message': 'No se encontraron reportes para esta cédula',
                'total_consultas': 0
            }), 200
        
        # Registrar consulta (sumar +1 al contador)
        cursor.execute("""
            SELECT id, count FROM consultas 
            WHERE user_id = %s
            ORDER BY id DESC LIMIT 1
        """, (request.usuario['id_user'],))
        
        consulta_existente = cursor.fetchone()
        
        if consulta_existente:
            nuevo_count = consulta_existente['count'] + 1
            cursor.execute("""
                UPDATE consultas 
                SET count = %s 
                WHERE id = %s
            """, (nuevo_count, consulta_existente['id']))
            total_consultas = nuevo_count
        else:
            cursor.execute("""
                INSERT INTO consultas (user_id, count)
                VALUES (%s, 1)
            """, (request.usuario['id_user'],))
            total_consultas = 1
        
        conn.commit()
        
        # Formatear respuesta
        reportes_formateados = []
        for r in reportes:
            reportes_formateados.append({
                'id': r['id'],
                'fecha_reporte': r['Fecha_Reporte'].strftime('%Y-%m-%d') if r['Fecha_Reporte'] else None,
                'numero_documento': r['Numero_Documento'],
                'nombres': r['Nombres'],
                'apellidos': r['Apellidos'],
                'fecha_cierre': r['Fecha_cierre'],
                'placa': r['Placa'],
                'valor_reporte': r['Valor_Reporte'],
                'descripcion': r['Descripcion_Reporte'],
                'vehiculo_afiliado': r['Vehiculo_afiliado'],
                'estado': r['Estado']
            })
        
        return jsonify({
            'success': True,
            'found': True,
            'total_reportes': len(reportes),
            'total_consultas': total_consultas,
            'reportes': reportes_formateados
        }), 200
        
    except Error as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 4: DESCARGAR PLANTILLA EXCEL ====================
@app.route('/api/plantilla-excel', methods=['GET'])
@verificar_usuario
def descargar_plantilla():
    """
    Descargar plantilla Excel para importación masiva
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
    responses:
      200:
        description: Archivo Excel
        content:
          application/vnd.openxmlformats-officedocument.spreadsheetml.sheet:
            schema:
              type: string
              format: binary
      401:
        description: No autorizado
      500:
        description: Error del servidor
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Reportes"
        
        headers = [
            'Fecha_Reporte', 'Numero_Documento', 'Nombres', 'Apellidos',
            'Fecha_cierre', 'Placa', 'Valor_Reporte', 'Descripcion_Reporte',
            'Vehiculo_afiliado', 'Estado'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        ejemplo = [
            '2024-01-15', '1234567890', 'JUAN', 'PEREZ GOMEZ',
            '', 'ABC123', '50000', 'REPORTE NEGATIVO POR TARIFAS',
            'ADMICARS', 'ACTIVA'
        ]
        
        for col, valor in enumerate(ejemplo, 1):
            ws.cell(row=2, column=col, value=valor)
        
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = max_length + 2
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'plantilla_reportes_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== SERVICIO 5: IMPORTAR DATOS MASIVOS DESDE EXCEL ====================
@app.route('/api/importar-excel', methods=['POST'])
@verificar_usuario
def importar_excel():
    """
    Importar datos masivos desde archivo Excel
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    consumes:
      - multipart/form-data
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
      - name: file
        in: formData
        type: file
        required: true
        description: Archivo Excel (.xlsx o .xls) con los datos a importar
    responses:
      200:
        description: Importación completada
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
            insertados:
              type: integer
            errores:
              type: integer
            detalle_errores:
              type: array
              items:
                type: string
      400:
        description: Archivo no válido
      401:
        description: No autorizado
      500:
        description: Error del servidor
    """
    if 'file' not in request.files:
        return jsonify({
            'success': False,
            'message': 'No se proporcionó archivo'
        }), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({
            'success': False,
            'message': 'Archivo vacío'
        }), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({
            'success': False,
            'message': 'Formato no válido. Use .xlsx o .xls'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        df = pd.read_excel(file)
        
        required_cols = ['Numero_Documento', 'Nombres', 'Apellidos', 'Placa']
        if not all(col in df.columns for col in required_cols):
            return jsonify({
                'success': False,
                'message': f'Columnas requeridas: {", ".join(required_cols)}'
            }), 400
        
        cursor = conn.cursor()
        insertados = 0
        errores = []
        
        for idx, row in df.iterrows():
            try:
                query = """
                    INSERT INTO personas (
                        Fecha_Reporte, Numero_Documento, Nombres, Apellidos,
                        Fecha_cierre, Placa, Valor_Reporte, Descripcion_Reporte,
                        Vehiculo_afiliado, Estado, Reportante_Nombres
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                
                valores = (
                    row.get('Fecha_Reporte', datetime.now().date()),
                    row['Numero_Documento'],
                    row['Nombres'],
                    row['Apellidos'],
                    row.get('Fecha_cierre', ''),
                    row['Placa'],
                    row.get('Valor_Reporte', 0),
                    row.get('Descripcion_Reporte', ''),
                    row.get('Vehiculo_afiliado', 'ADMICARS'),
                    row.get('Estado', 'ACTIVA'),
                    request.usuario['id_user']
                )
                
                cursor.execute(query, valores)
                insertados += 1
                
            except Exception as e:
                errores.append(f"Fila {idx + 2}: {str(e)}")
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': f'Importación completada',
            'insertados': insertados,
            'errores': len(errores),
            'detalle_errores': errores[:10]
        }), 200
        
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 6: CREAR REPORTE INDIVIDUAL ====================
@app.route('/api/personas', methods=['POST'])
@verificar_usuario
def crear_persona():
    """
    Crear un reporte individual
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
      - name: body
        in: body
        required: true
        schema:
          type: object
          required:
            - numero_documento
            - nombres
            - apellidos
            - placa
          properties:
            numero_documento:
              type: string
              example: "1234567890"
            nombres:
              type: string
              example: "JUAN"
            apellidos:
              type: string
              example: "PEREZ GOMEZ"
            placa:
              type: string
              example: "ABC123"
            valor_reporte:
              type: integer
              example: 50000
            descripcion:
              type: string
              example: "REPORTE NEGATIVO POR TARIFAS"
            vehiculo_afiliado:
              type: string
              example: "ADMICARS"
            estado:
              type: string
              example: "ACTIVA"
    responses:
      201:
        description: Reporte creado exitosamente
      400:
        description: Parámetros incorrectos
      401:
        description: No autorizado
      500:
        description: Error del servidor
    """
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Error al procesar JSON: ' + str(e)
        }), 400
    
    if not data:
        return jsonify({
            'success': False,
            'message': 'No se recibió ningún dato JSON'
        }), 400
    
    required = ['numero_documento', 'nombres', 'apellidos', 'placa']
    if not all(field in data for field in required):
        return jsonify({
            'success': False,
            'message': 'Campos requeridos: numero_documento, nombres, apellidos, placa'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor()
        
        query = """
            INSERT INTO personas (
                Fecha_Reporte, Numero_Documento, Nombres, Apellidos,
                Placa, Valor_Reporte, Descripcion_Reporte,
                Vehiculo_afiliado, Estado, Reportante_Nombres
            ) VALUES (NOW(), %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            data['numero_documento'],
            data['nombres'].upper(),
            data['apellidos'].upper(),
            data['placa'].upper(),
            data.get('valor_reporte', 0),
            data.get('descripcion', ''),
            data.get('vehiculo_afiliado', 'ADMICARS'),
            data.get('estado', 'ACTIVA'),
            request.usuario['id_user']
        ))
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Reporte creado exitosamente',
            'id': cursor.lastrowid
        }), 201
        
    except Error as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 7: EDITAR REPORTE ====================
@app.route('/api/personas/<int:id>', methods=['PUT'])
@verificar_usuario
def editar_persona(id):
    """
    Editar un reporte (solo si lo creó el usuario o es admin)
    ---
    tags:
      - Reportes
    security:
      - CelularAuth: []
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
      - name: id
        in: path
        type: integer
        required: true
        description: ID del reporte a editar
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            nombres:
              type: string
            apellidos:
              type: string
            placa:
              type: string
            valor_reporte:
              type: integer
            descripcion_reporte:
              type: string
            estado:
              type: string
            fecha_cierre:
              type: string
    responses:
      200:
        description: Reporte actualizado exitosamente
      400:
        description: No hay campos para actualizar
      401:
        description: No autorizado
      403:
        description: Sin permisos para editar
      404:
        description: Reporte no encontrado
      500:
        description: Error del servidor
    """
    try:
        data = request.get_json(force=True) or {}
    except Exception as e:
        return jsonify({
            'success': False,
            'message': 'Error al procesar JSON: ' + str(e)
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT Reportante_Nombres FROM personas WHERE id = %s
        """, (id,))
        
        reporte = cursor.fetchone()
        
        if not reporte:
            return jsonify({
                'success': False,
                'message': 'Reporte no encontrado'
            }), 404
        
        es_admin = request.usuario['rol'] == 'admin'
        es_creador = str(reporte['Reportante_Nombres']) == str(request.usuario['id_user'])
        
        if not (es_admin or es_creador):
            return jsonify({
                'success': False,
                'message': 'No tiene permisos para editar este reporte'
            }), 403
        
        campos_permitidos = [
            'Nombres', 'Apellidos', 'Placa', 'Valor_Reporte',
            'Descripcion_Reporte', 'Estado', 'Fecha_cierre'
        ]
        
        campos_actualizar = []
        valores = []
        
        for campo in campos_permitidos:
            campo_lower = campo.lower().replace('_', '')
            for key in data.keys():
                if key.lower().replace('_', '') == campo_lower:
                    campos_actualizar.append(f"{campo} = %s")
                    valores.append(data[key])
                    break
        
        if not campos_actualizar:
            return jsonify({
                'success': False,
                'message': 'No hay campos para actualizar'
            }), 400
        
        valores.append(id)
        query = f"UPDATE personas SET {', '.join(campos_actualizar)} WHERE id = %s"
        
        cursor.execute(query, valores)
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Reporte actualizado exitosamente'
        }), 200
        
    except Error as e:
        conn.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 8: OBTENER ESTADO DE CONVERSACIÓN ====================
@app.route('/api/estado-usuario/<celular>', methods=['GET'])
def obtener_estado_usuario(celular):
    """
    Obtener el estado actual de conversación de un usuario
    ---
    tags:
      - Estado de Conversación
    parameters:
      - name: celular
        in: path
        type: string
        required: true
        description: Número de celular del usuario
        example: "3007471199"
    responses:
      200:
        description: Estado encontrado o no encontrado
        schema:
          type: object
          properties:
            success:
              type: boolean
            exists:
              type: boolean
            estado:
              type: object
              properties:
                celular:
                  type: string
                estado:
                  type: string
                opcion:
                  type: integer
                updated_at:
                  type: string
      500:
        description: Error del servidor
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Error conectando a la base de datos'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT celular, estado, opcion, updated_at
            FROM user_state
            WHERE celular = %s
        """, (celular,))
        
        estado = cursor.fetchone()
        
        if estado:
            return jsonify({
                'success': True,
                'exists': True,
                'estado': {
                    'celular': estado['celular'],
                    'estado': estado['estado'],
                    'opcion': estado['opcion'],
                    'updated_at': estado['updated_at'].isoformat() if estado['updated_at'] else None
                }
            }), 200
        else:
            return jsonify({
                'success': True,
                'exists': False,
                'estado': None
            }), 200
    
    except Error as e:
        return jsonify({
            'success': False,
            'message': f'Error consultando estado: {str(e)}'
        }), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 9: GUARDAR ESTADO DE CONVERSACIÓN ====================
@app.route('/api/estado-usuario', methods=['POST'])
def guardar_estado_usuario():
    """
    Guardar o actualizar el estado de conversación de un usuario
    ---
    tags:
      - Estado de Conversación
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          required:
            - celular
            - estado
          properties:
            celular:
              type: string
              example: "3007471199"
            estado:
              type: string
              example: "esperando_cedula"
            opcion:
              type: integer
              example: 1
    responses:
      200:
        description: Estado guardado exitosamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
      400:
        description: Parámetros incorrectos
      500:
        description: Error del servidor
    """
    try:
        data = request.get_json(force=True)
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error parseando JSON: {str(e)}'
        }), 400
    
    if not data:
        return jsonify({
            'success': False,
            'message': 'Cuerpo de la solicitud vacío'
        }), 400
    
    celular = data.get('celular')
    estado = data.get('estado')
    opcion = data.get('opcion')
    
    if not celular or not estado:
        return jsonify({
            'success': False,
            'message': 'Faltan campos requeridos: celular y estado'
        }), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Error conectando a la base de datos'
        }), 500
    
    try:
        cursor = conn.cursor()
        
        # Usar INSERT ... ON DUPLICATE KEY UPDATE para insertar o actualizar
        cursor.execute("""
            INSERT INTO user_state (celular, estado, opcion, updated_at)
            VALUES (%s, %s, %s, NOW())
            ON DUPLICATE KEY UPDATE
                estado = VALUES(estado),
                opcion = VALUES(opcion),
                updated_at = NOW()
        """, (celular, estado, opcion))
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Estado guardado exitosamente'
        }), 200
    
    except Error as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Error guardando estado: {str(e)}'
        }), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 10: ELIMINAR ESTADO DE CONVERSACIÓN ====================
@app.route('/api/estado-usuario/<celular>', methods=['DELETE'])
def eliminar_estado_usuario(celular):
    """
    Eliminar el estado de conversación de un usuario (limpiar estado)
    ---
    tags:
      - Estado de Conversación
    parameters:
      - name: celular
        in: path
        type: string
        required: true
        description: Número de celular del usuario
        example: "3007471199"
    responses:
      200:
        description: Estado eliminado exitosamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
      500:
        description: Error del servidor
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Error conectando a la base de datos'
        }), 500
    
    try:
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM user_state
            WHERE celular = %s
        """, (celular,))
        
        conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Estado eliminado exitosamente'
        }), 200
    
    except Error as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Error eliminando estado: {str(e)}'
        }), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO 11: BLOQUEAR/DESBLOQUEAR USUARIO ====================
@app.route('/api/usuarios/<celular>/bloquear', methods=['PUT'])
def bloquear_usuario(celular):
    """
    Bloquear o desbloquear un usuario por celular
    ---
    tags:
      - Usuarios
    parameters:
      - name: celular
        in: path
        type: string
        required: true
        description: Número de celular del usuario a bloquear
        example: "3001234567"
      - name: body
        in: body
        required: false
        schema:
          type: object
          properties:
            bloquear:
              type: boolean
              example: true
              description: true para bloquear, false para desbloquear (por defecto true)
    responses:
      200:
        description: Usuario bloqueado/desbloqueado exitosamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
      404:
        description: Usuario no encontrado
      500:
        description: Error del servidor
    """
    try:
        data = request.get_json(force=True) or {}
        bloquear = data.get('bloquear', True)  # Por defecto bloquear
    except Exception:
        bloquear = True
    
    conn = get_db_connection()
    if not conn:
        return jsonify({
            'success': False,
            'message': 'Error conectando a la base de datos'
        }), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Verificar si el usuario existe
        cursor.execute("""
            SELECT id_user, username, isactive 
            FROM users 
            WHERE Celular = %s
        """, (celular,))
        
        usuario = cursor.fetchone()
        
        if not usuario:
            return jsonify({
                'success': False,
                'message': f'Usuario con celular {celular} no encontrado'
            }), 404
        
        # Actualizar el estado isactive
        nuevo_estado = 0 if bloquear else 1
        cursor.execute("""
            UPDATE users 
            SET isactive = %s 
            WHERE Celular = %s
        """, (nuevo_estado, celular))
        
        conn.commit()
        
        accion = 'bloqueado' if bloquear else 'desbloqueado'
        return jsonify({
            'success': True,
            'message': f'Usuario {usuario["username"]} {accion} exitosamente'
        }), 200
    
    except Error as e:
        conn.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al bloquear usuario: {str(e)}'
        }), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== SERVICIO EXTRA: ESTADÍSTICAS DE USUARIO ====================
@app.route('/api/estadisticas', methods=['GET'])
@verificar_usuario
def obtener_estadisticas():
    """
    Obtener estadísticas del usuario actual
    ---
    tags:
      - Estadísticas
    security:
      - CelularAuth: []
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Número de celular del usuario autenticado
        example: "3007471199"
    responses:
      200:
        description: Estadísticas del usuario
        schema:
          type: object
          properties:
            success:
              type: boolean
            usuario:
              type: object
            total_consultas:
              type: integer
            reportes_creados:
              type: integer
      401:
        description: No autorizado
      500:
        description: Error del servidor
    """
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Error de conexión'}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # Total de consultas
        cursor.execute("""
            SELECT COALESCE(SUM(count), 0) as total
            FROM consultas
            WHERE user_id = %s
        """, (request.usuario['id_user'],))
        total_consultas = cursor.fetchone()['total']
        
        # Reportes creados por el usuario
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM personas
            WHERE Reportante_Nombres = %s
        """, (request.usuario['id_user'],))
        reportes_creados = cursor.fetchone()['total']
        
        return jsonify({
            'success': True,
            'usuario': {
                'id': request.usuario['id_user'],
                'nombre': request.usuario['nombres'],
                'email': request.usuario['username'],
                'rol': request.usuario['rol']
            },
            'total_consultas': total_consultas,
            'reportes_creados': reportes_creados
        }), 200
        
    except Error as e:
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

# ==================== RUTA DE PRUEBA ====================
@app.route('/api/health', methods=['GET', 'OPTIONS'])
def health_check():
    """
    Verificar que la API está funcionando
    ---
    tags:
      - Sistema
    responses:
      200:
        description: API funcionando correctamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
            timestamp:
              type: string
            database:
              type: string
    """
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization,Accept")
        response.headers.add('Access-Control-Allow-Methods', "GET,PUT,POST,DELETE,OPTIONS")
        response.headers.add('Access-Control-Max-Age', "3600")
        return response
    
    try:
        conn = get_db_connection()
        db_status = "Conectada" if conn else "Error de conexión"
        if conn and conn.is_connected():
            conn.close()
        
        response = jsonify({
            'success': True,
            'message': 'API InfoTaxi funcionando correctamente',
            'timestamp': datetime.now().isoformat(),
            'database': db_status,
            'swagger_docs': '/apidocs/'
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization,Accept')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        response.headers.add('Access-Control-Max-Age', '3600')
        return response, 200
    except Exception as e:
        response = jsonify({
            'success': False,
            'message': f'Error en health check: {str(e)}'
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        response.headers.add('Access-Control-Allow-Headers', 'Content-Type,X-User-Celular,Authorization,Accept')
        response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
        return response, 500

# ==================== ENDPOINT: GENERAR TOKEN PARA CARGA MASIVA ====================
@app.route('/api/generar-token-carga', methods=['POST', 'OPTIONS'])
def generar_token_carga():
    """
    Genera un token temporal para carga masiva asociado a un usuario
    ---
    tags:
      - Excel
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Celular del usuario
    responses:
      200:
        description: Token generado exitosamente
        schema:
          type: object
          properties:
            success:
              type: boolean
            token:
              type: string
            url:
              type: string
            expira_en:
              type: string
      500:
        description: Error al generar token
    """
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization,Accept")
        response.headers.add('Access-Control-Allow-Methods', "POST,OPTIONS")
        return response
    
    celular = request.headers.get('X-User-Celular')
    
    if not celular:
        return jsonify({
            'success': False,
            'message': 'Celular requerido en headers'
        }), 401
    
    try:
        conn = get_db_connection()
        if not conn:
            return jsonify({'success': False, 'message': 'Error de conexión a BD'}), 500
        
        cursor = conn.cursor(dictionary=True)
        
        # Verificar que el usuario existe
        cursor.execute("SELECT id_user, nombres FROM users WHERE Celular = %s AND isactive = 1", (celular,))
        usuario = cursor.fetchone()
        
        if not usuario:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'message': 'Usuario no encontrado o inactivo'
            }), 403
        
        # Generar token único
        token = secrets.token_urlsafe(32)
        
        # Fecha de expiración (24 horas)
        expiracion = datetime.now() + timedelta(hours=24)
        
        # Guardar en user_state (reutilizamos la tabla)
        cursor.execute("""
            INSERT INTO user_state (celular, estado, opcion, updated_at) 
            VALUES (%s, %s, 4, %s)
            ON DUPLICATE KEY UPDATE estado = %s, opcion = 4, updated_at = %s
        """, (celular, f"token_carga:{token}:{expiracion.isoformat()}", datetime.now(), f"token_carga:{token}:{expiracion.isoformat()}", datetime.now()))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        # Construir URL completa
        base_url = request.host_url.rstrip('/')
        upload_url = f"{base_url}/carga-masiva/{token}"
        
        response = jsonify({
            'success': True,
            'token': token,
            'url': upload_url,
            'expira_en': '24 horas',
            'mensaje': f'Link generado para {usuario["nombres"]}'
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al generar token: {str(e)}'
        }), 500

# ==================== PÁGINA WEB: CARGA MASIVA ====================
@app.route('/carga-masiva/<token>', methods=['GET'])
def pagina_carga_masiva(token):
    """Página web para descargar plantilla y subir archivo"""
    
    # Validar token
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT celular, estado, updated_at FROM user_state WHERE estado LIKE %s", (f"token_carga:{token}:%",))
        token_data = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not token_data:
            return """
            <!DOCTYPE html>
            <html><head><meta charset="UTF-8"><title>Token Inválido</title></head>
            <body style="font-family:Arial;text-align:center;padding:50px;">
                <h1>❌ Token Inválido</h1>
                <p>Este link no existe o ya expiró.</p>
            </body></html>
            """, 404
        
        # Extraer fecha de expiración del estado
        estado_parts = token_data['estado'].split(':')
        if len(estado_parts) < 3:
            return "Token mal formado", 400
        
        fecha_expiracion_str = estado_parts[2]
        fecha_expiracion = datetime.fromisoformat(fecha_expiracion_str)
        
        # Verificar si expiró
        if datetime.now() > fecha_expiracion:
            return """
            <!DOCTYPE html>
            <html><head><meta charset="UTF-8"><title>Token Expirado</title></head>
            <body style="font-family:Arial;text-align:center;padding:50px;">
                <h1>⏰ Token Expirado</h1>
                <p>Este link ya caducó. Solicita uno nuevo desde WhatsApp.</p>
            </body></html>
            """, 403
        
        # Obtener datos del usuario
        celular = token_data['celular']
        
        html = f"""
        <!DOCTYPE html>
        <html lang="es">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>InfoTaxi - Carga Masiva</title>
            <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
            <style>
                * {{ margin: 0; padding: 0; box-sizing: border-box; }}
                body {{ 
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; 
                    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                    min-height: 100vh;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    padding: 20px;
                }}
                .container {{
                    background: white;
                    padding: 40px;
                    border-radius: 15px;
                    box-shadow: 0 10px 30px rgba(0,0,0,0.2);
                    max-width: 600px;
                    width: 100%;
                }}
                h1 {{ color: #2c3e50; margin-bottom: 10px; text-align: center; }}
                .subtitle {{ text-align: center; color: #7f8c8d; margin-bottom: 30px; }}
                .section {{
                    background: #f8f9fa;
                    padding: 25px;
                    border-radius: 10px;
                    margin-bottom: 25px;
                    border-left: 4px solid #f39c12;
                }}
                .section h2 {{
                    color: #2c3e50;
                    font-size: 20px;
                    margin-bottom: 15px;
                }}
                .btn {{
                    display: inline-block;
                    padding: 15px 30px;
                    background: #27ae60;
                    color: white;
                    text-decoration: none;
                    border-radius: 8px;
                    font-weight: 600;
                    transition: all 0.3s;
                    border: none;
                    cursor: pointer;
                    width: 100%;
                    font-size: 16px;
                    margin-top: 10px;
                }}
                .btn:hover {{ background: #229954; transform: translateY(-2px); }}
                .btn-upload {{ background: #f39c12; }}
                .btn-upload:hover {{ background: #e67e22; }}
                .file-input {{
                    display: none;
                }}
                .file-label {{
                    display: block;
                    padding: 40px;
                    border: 3px dashed #ddd;
                    border-radius: 10px;
                    text-align: center;
                    cursor: pointer;
                    transition: all 0.3s;
                    margin: 15px 0;
                }}
                .file-label:hover {{ border-color: #f39c12; background: #fef9e7; }}
                .file-label i {{ font-size: 48px; color: #f39c12; display: block; margin-bottom: 10px; }}
                .result {{
                    margin-top: 20px;
                    padding: 15px;
                    border-radius: 8px;
                    display: none;
                }}
                .result.success {{ background: #d4edda; color: #155724; border: 1px solid #c3e6cb; }}
                .result.error {{ background: #f8d7da; color: #721c24; border: 1px solid #f5c6cb; }}
                .loading {{
                    text-align: center;
                    padding: 20px;
                    display: none;
                }}
                .loading i {{ font-size: 48px; color: #f39c12; animation: spin 1s linear infinite; }}
                @keyframes spin {{ from {{ transform: rotate(0deg); }} to {{ transform: rotate(360deg); }} }}
                .info {{ background: #d1ecf1; padding: 15px; border-radius: 8px; border-left: 4px solid #17a2b8; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1><i class="fas fa-taxi"></i> InfoTaxi</h1>
                <p class="subtitle">Carga Masiva de Reportes</p>
                
                <div class="info">
                    <i class="fas fa-info-circle"></i> 
                    <strong>Link personal de carga</strong><br>
                    Usuario: {celular}<br>
                    Expira: {fecha_expiracion.strftime('%d/%m/%Y %H:%M')}
                </div>
                
                <div class="section">
                    <h2><i class="fas fa-download"></i> Paso 1: Descargar Plantilla</h2>
                    <p>Descarga la plantilla Excel con el formato correcto:</p>
                    <a href="/api/plantilla-excel-token/{token}" class="btn" download>
                        <i class="fas fa-file-excel"></i> Descargar Plantilla Excel
                    </a>
                </div>
                
                <div class="section">
                    <h2><i class="fas fa-upload"></i> Paso 2: Subir Archivo Completado</h2>
                    <p>Selecciona tu archivo Excel con los reportes:</p>
                    <form id="uploadForm" enctype="multipart/form-data">
                        <input type="file" id="excelFile" name="file" accept=".xlsx" class="file-input" onchange="updateFileName()">
                        <label for="excelFile" class="file-label">
                            <i class="fas fa-cloud-upload-alt"></i>
                            <span id="fileName">Click para seleccionar archivo .xlsx</span>
                        </label>
                        <button type="submit" class="btn btn-upload">
                            <i class="fas fa-upload"></i> Importar Reportes
                        </button>
                    </form>
                    
                    <div class="loading" id="loading">
                        <i class="fas fa-spinner"></i>
                        <p>Procesando archivo...</p>
                    </div>
                    
                    <div class="result" id="result"></div>
                </div>
            </div>
            
            <script>
                const token = '{token}';
                const celular = '{celular}';
                
                function updateFileName() {{
                    const input = document.getElementById('excelFile');
                    const label = document.getElementById('fileName');
                    if (input.files.length > 0) {{
                        label.textContent = input.files[0].name;
                    }}
                }}
                
                document.getElementById('uploadForm').addEventListener('submit', async (e) => {{
                    e.preventDefault();
                    
                    const fileInput = document.getElementById('excelFile');
                    const file = fileInput.files[0];
                    
                    if (!file) {{
                        showResult('Por favor selecciona un archivo', 'error');
                        return;
                    }}
                    
                    const formData = new FormData();
                    formData.append('file', file);
                    
                    document.getElementById('loading').style.display = 'block';
                    document.getElementById('result').style.display = 'none';
                    
                    try {{
                        const response = await fetch(`/api/importar-excel-token/${{token}}`, {{
                            method: 'POST',
                            body: formData
                        }});
                        
                        const data = await response.json();
                        
                        document.getElementById('loading').style.display = 'none';
                        
                        if (data.success) {{
                            let message = `✅ Importación completada\\n`;
                            message += `Total: ${{data.total_filas}}\\n`;
                            message += `Importados: ${{data.importados}}\\n`;
                            message += `Errores: ${{data.errores}}`;
                            
                            if (data.detalles && data.detalles.length > 0) {{
                                message += '\\n\\nDetalles:\\n';
                                data.detalles.slice(0, 10).forEach(det => {{
                                    const icon = det.status === 'success' ? '✅' : '❌';
                                    message += `${{icon}} Fila ${{det.fila}}: ${{det.mensaje}}\\n`;
                                }});
                            }}
                            
                            showResult(message, 'success');
                            fileInput.value = '';
                            document.getElementById('fileName').textContent = 'Click para seleccionar archivo .xlsx';
                        }} else {{
                            showResult('❌ Error: ' + (data.message || 'Error desconocido'), 'error');
                        }}
                    }} catch (error) {{
                        document.getElementById('loading').style.display = 'none';
                        showResult('❌ Error de conexión: ' + error.message, 'error');
                    }}
                }});
                
                function showResult(message, type) {{
                    const resultDiv = document.getElementById('result');
                    resultDiv.textContent = message;
                    resultDiv.className = 'result ' + type;
                    resultDiv.style.display = 'block';
                }}
            </script>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        return f"Error al cargar la página: {str(e)}", 500

# ==================== ENDPOINT: DESCARGAR PLANTILLA CON TOKEN ====================
@app.route('/api/plantilla-excel-token/<token>', methods=['GET'])
def descargar_plantilla_con_token(token):
    """Descarga plantilla Excel validando el token"""
    
    # Validar token
    conn = get_db_connection()
    if not conn:
        return jsonify({{'success': False, 'message': 'Error de conexión'}}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT celular FROM user_state WHERE estado LIKE %s", (f"token_carga:{token}:%",))
        token_data = cursor.fetchone()
        cursor.close()
        conn.close()
        
        if not token_data:
            return jsonify({{'success': False, 'message': 'Token inválido'}}), 404
        
        # Crear plantilla Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Reportes"
        
        # Headers según la imagen proporcionada
        headers = [
            'Documento Conductor',
            'Nombre Conductor',
            'Apellidos Conductor',
            'Fecha Inicio Reporte',
            'Placa Vehiculo',
            'Valor del Reporte',
            'Descripcion del Reporte',
            'Vehiculo Afiliado'
        ]
        
        # Estilo para el header
        header_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 20  # Documento Conductor
        ws.column_dimensions['B'].width = 20  # Nombre Conductor
        ws.column_dimensions['C'].width = 22  # Apellidos Conductor
        ws.column_dimensions['D'].width = 20  # Fecha Inicio Reporte
        ws.column_dimensions['E'].width = 15  # Placa Vehiculo
        ws.column_dimensions['F'].width = 18  # Valor del Reporte
        ws.column_dimensions['G'].width = 40  # Descripcion del Reporte
        ws.column_dimensions['H'].width = 18  # Vehiculo Afiliado
        
        # Agregar fila de ejemplo
        ws.append([
            '123456789',
            'Juan',
            'Pérez',
            '2026-01-15',
            'ABC123',
            '50000',
            'Descripción del reporte',
            'SI'
        ])
        
        # Guardar en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        filename = f"Plantilla_Importacion.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({{'success': False, 'message': f'Error al generar plantilla: {{str(e)}}' }}), 500

# ==================== ENDPOINT: IMPORTAR EXCEL CON TOKEN ====================
@app.route('/api/importar-excel-token/<token>', methods=['POST'])
def importar_excel_con_token(token):
    """Importa reportes validando el token"""
    
    # Validar token
    conn = get_db_connection()
    if not conn:
        return jsonify({{'success': False, 'message': 'Error de conexión'}}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT celular FROM user_state WHERE estado LIKE %s", (f"token_carga:{token}:%",))
        token_data = cursor.fetchone()
        
        if not token_data:
            cursor.close()
            conn.close()
            return jsonify({{'success': False, 'message': 'Token inválido'}}), 404
        
        celular = token_data['celular']
        
        # Obtener id_user
        cursor.execute("SELECT id_user FROM users WHERE Celular = %s", (celular,))
        user_result = cursor.fetchone()
        
        if not user_result:
            cursor.close()
            conn.close()
            return jsonify({{'success': False, 'message': 'Usuario no encontrado'}}), 400
        
        id_user = user_result['id_user']
        
        # Verificar archivo
        if 'file' not in request.files:
            cursor.close()
            conn.close()
            return jsonify({{'success': False, 'message': 'No se envió ningún archivo'}}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            cursor.close()
            conn.close()
            return jsonify({{'success': False, 'message': 'No se seleccionó ningún archivo'}}), 400
        
        if not file.filename.endswith('.xlsx'):
            cursor.close()
            conn.close()
            return jsonify({{'success': False, 'message': 'El archivo debe ser formato .xlsx'}}), 400
        
        # Leer Excel
        df = pd.read_excel(file, engine='openpyxl')
        
        # Validar columnas según la plantilla
        columnas_requeridas = [
            'Documento Conductor', 'Nombre Conductor', 'Apellidos Conductor',
            'Placa Vehiculo', 'Valor del Reporte', 'Descripcion del Reporte'
        ]
        
        for col in columnas_requeridas:
            if col not in df.columns:
                cursor.close()
                conn.close()
                return jsonify({{
                    'success': False,
                    'message': f'Falta la columna requerida: {{col}}'
                }}), 400
        
        # Procesar filas
        total_filas = len(df)
        importados = 0
        errores = 0
        detalles = []
        
        for index, row in df.iterrows():
            fila_num = index + 2
            
            try:
                # Validar campos
                if pd.isna(row['Documento Conductor']) or pd.isna(row['Nombre Conductor']) or pd.isna(row['Apellidos Conductor']):
                    detalles.append({{
                        'fila': fila_num,
                        'status': 'error',
                        'mensaje': 'Campos obligatorios vacíos'
                    }})
                    errores += 1
                    continue
                
                # Extraer datos
                numero_doc = str(row['Documento Conductor']).strip()
                nombres = str(row['Nombre Conductor']).strip()
                apellidos = str(row['Apellidos Conductor']).strip()
                placa = str(row['Placa Vehiculo']).strip() if not pd.isna(row['Placa Vehiculo']) else ''
                valor = str(row['Valor del Reporte']).strip() if not pd.isna(row['Valor del Reporte']) else '0'
                descripcion = str(row['Descripcion del Reporte']).strip() if not pd.isna(row['Descripcion del Reporte']) else ''
                vehiculo_afiliado = str(row['Vehiculo Afiliado']).strip().upper() if not pd.isna(row['Vehiculo Afiliado']) else 'No'
                
                # Insertar
                query = """
                    INSERT INTO personas 
                    (Fecha_Reporte, Numero_Documento, Nombres, Apellidos, 
                     Placa, Valor_Reporte, Descripcion_Reporte, 
                     Vehiculo_afiliado, Estado, Reportante_Nombres)
                    VALUES (NOW(), %s, %s, %s, %s, %s, %s, %s, 'Activo', %s)
                """
                
                cursor.execute(query, (
                    numero_doc, nombres, apellidos, 
                    placa, valor, descripcion, vehiculo_afiliado, id_user
                ))
                
                importados += 1
                detalles.append({{
                    'fila': fila_num,
                    'status': 'success',
                    'mensaje': f'Reporte importado: {{nombres}} {{apellidos}}'
                }})
                
            except Exception as e:
                errores += 1
                detalles.append({{
                    'fila': fila_num,
                    'status': 'error',
                    'mensaje': f'Error: {{str(e)}}'
                }})
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({{
            'success': True,
            'message': f'Importación completada: {{importados}} exitosos, {{errores}} errores',
            'total_filas': total_filas,
            'importados': importados,
            'errores': errores,
            'detalles': detalles
        }}), 200
        
    except Exception as e:
        return jsonify({{
            'success': False,
            'message': f'Error al procesar el archivo: {{str(e)}}'
        }}), 500

# ==================== ENDPOINT: DESCARGAR PLANTILLA EXCEL (ORIGINAL) ====================
@app.route('/api/plantilla-excel', methods=['GET', 'OPTIONS'])
def descargar_plantilla_excel():
    """
    Descarga plantilla Excel para carga masiva de reportes
    ---
    tags:
      - Excel
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: false
        description: Celular del usuario solicitante
    responses:
      200:
        description: Archivo Excel descargado exitosamente
        schema:
          type: file
      500:
        description: Error al generar la plantilla
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
    """
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization,Accept")
        response.headers.add('Access-Control-Allow-Methods', "GET,OPTIONS")
        return response
    
    try:
        # Crear un nuevo workbook con openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla Reportes"
        
        # Definir los headers
        headers = [
            'Numero_Documento',
            'Nombres', 
            'Apellidos',
            'Placa',
            'Valor_Reporte',
            'Descripcion_Reporte'
        ]
        
        # Estilo para el header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers en la primera fila
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 18  # Numero_Documento
        ws.column_dimensions['B'].width = 25  # Nombres
        ws.column_dimensions['C'].width = 25  # Apellidos
        ws.column_dimensions['D'].width = 12  # Placa
        ws.column_dimensions['E'].width = 15  # Valor_Reporte
        ws.column_dimensions['F'].width = 50  # Descripcion_Reporte
        
        # Agregar una fila de ejemplo
        ws.append([
            '1234567890',
            'Juan Carlos',
            'Pérez García',
            'ABC123',
            '50000',
            'Servicio mal prestado, conductor grosero'
        ])
        
        # Guardar el archivo en memoria
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Nombre del archivo con fecha
        filename = f"plantilla_reportes_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        response = send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response
        
    except Exception as e:
        response = jsonify({
            'success': False,
            'message': f'Error al generar plantilla: {str(e)}'
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 500

# ==================== ENDPOINT: IMPORTAR EXCEL ====================
@app.route('/api/importar-excel', methods=['POST', 'OPTIONS'])
@require_celular
def importar_excel():
    """
    Importa reportes desde archivo Excel
    ---
    tags:
      - Excel
    consumes:
      - multipart/form-data
    parameters:
      - name: X-User-Celular
        in: header
        type: string
        required: true
        description: Celular del usuario que importa (debe existir en users)
      - name: file
        in: formData
        type: file
        required: true
        description: Archivo Excel con los reportes (.xlsx)
    responses:
      200:
        description: Importación completada
        schema:
          type: object
          properties:
            success:
              type: boolean
            message:
              type: string
            total_filas:
              type: integer
            importados:
              type: integer
            errores:
              type: integer
            detalles:
              type: array
              items:
                type: object
                properties:
                  fila:
                    type: integer
                  status:
                    type: string
                  mensaje:
                    type: string
      400:
        description: Error en la solicitud
      500:
        description: Error en el servidor
    """
    if request.method == 'OPTIONS':
        response = jsonify({'status': 'ok'})
        response.headers.add("Access-Control-Allow-Origin", "*")
        response.headers.add('Access-Control-Allow-Headers', "Content-Type,X-User-Celular,Authorization,Accept")
        response.headers.add('Access-Control-Allow-Methods', "POST,OPTIONS")
        return response
    
    celular = request.headers.get('X-User-Celular')
    
    # Verificar que se envió un archivo
    if 'file' not in request.files:
        return jsonify({
            'success': False,
            'message': 'No se envió ningún archivo'
        }), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({
            'success': False,
            'message': 'No se seleccionó ningún archivo'
        }), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({
            'success': False,
            'message': 'El archivo debe ser formato .xlsx'
        }), 400
    
    try:
        # Leer el archivo Excel
        df = pd.read_excel(file, engine='openpyxl')
        
        # Validar que tenga las columnas requeridas
        columnas_requeridas = [
            'Numero_Documento', 'Nombres', 'Apellidos', 
            'Placa', 'Valor_Reporte', 'Descripcion_Reporte'
        ]
        
        for col in columnas_requeridas:
            if col not in df.columns:
                return jsonify({
                    'success': False,
                    'message': f'Falta la columna requerida: {col}'
                }), 400
        
        conn = get_db_connection()
        if not conn:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
        
        cursor = conn.cursor()
        
        # Obtener el id_user del celular
        cursor.execute("SELECT id_user FROM users WHERE Celular = %s", (celular,))
        user_result = cursor.fetchone()
        
        if not user_result:
            cursor.close()
            conn.close()
            return jsonify({
                'success': False,
                'message': 'Usuario no encontrado'
            }), 400
        
        id_user = user_result[0]
        
        # Procesar cada fila
        total_filas = len(df)
        importados = 0
        errores = 0
        detalles = []
        
        for index, row in df.iterrows():
            fila_num = index + 2  # +2 porque Excel empieza en 1 y tiene header
            
            try:
                # Validar que los campos requeridos no estén vacíos
                if pd.isna(row['Numero_Documento']) or pd.isna(row['Nombres']) or pd.isna(row['Apellidos']):
                    detalles.append({
                        'fila': fila_num,
                        'status': 'error',
                        'mensaje': 'Campos obligatorios vacíos (Documento, Nombres, Apellidos)'
                    })
                    errores += 1
                    continue
                
                # Preparar los valores
                numero_doc = str(row['Numero_Documento']).strip()
                nombres = str(row['Nombres']).strip()
                apellidos = str(row['Apellidos']).strip()
                placa = str(row['Placa']).strip() if not pd.isna(row['Placa']) else ''
                valor = str(row['Valor_Reporte']).strip() if not pd.isna(row['Valor_Reporte']) else '0'
                descripcion = str(row['Descripcion_Reporte']).strip() if not pd.isna(row['Descripcion_Reporte']) else ''
                
                # Insertar en la base de datos
                query = """
                    INSERT INTO personas 
                    (Fecha_Reporte, Numero_Documento, Nombres, Apellidos, 
                     Placa, Valor_Reporte, Descripcion_Reporte, 
                     Vehiculo_afiliado, Estado, Reportante_Nombres)
                    VALUES (NOW(), %s, %s, %s, %s, %s, %s, 'No', 'Activo', %s)
                """
                
                cursor.execute(query, (
                    numero_doc, nombres, apellidos, 
                    placa, valor, descripcion, id_user
                ))
                
                importados += 1
                detalles.append({
                    'fila': fila_num,
                    'status': 'success',
                    'mensaje': f'Reporte importado: {nombres} {apellidos}'
                })
                
            except Exception as e:
                errores += 1
                detalles.append({
                    'fila': fila_num,
                    'status': 'error',
                    'mensaje': f'Error al importar: {str(e)}'
                })
        
        conn.commit()
        cursor.close()
        conn.close()
        
        response = jsonify({
            'success': True,
            'message': f'Importación completada: {importados} exitosos, {errores} errores',
            'total_filas': total_filas,
            'importados': importados,
            'errores': errores,
            'detalles': detalles
        })
        response.headers.add('Access-Control-Allow-Origin', '*')
        return response, 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al procesar el archivo: {str(e)}'
        }), 500

# ==================== RUTA PRINCIPAL ====================
@app.route('/', methods=['GET'])
def index():
    """
    Página de inicio - Redirige a documentación Swagger
    ---
    tags:
      - Sistema
    responses:
      200:
        description: Información de la API
    """
    return jsonify({
        'message': 'API InfoTaxi',
        'version': '1.0.0',
        'documentacion': '/apidocs/',
        'endpoints': {
            'health': '/api/health',
            'verificar_usuario': '/api/verificar-usuario',
            'crear_usuario': '/api/usuarios',
            'consultar_persona': '/api/personas/<cedula>',
            'plantilla_excel': '/api/plantilla-excel',
            'importar_excel': '/api/importar-excel',
            'crear_reporte': '/api/personas',
            'editar_reporte': '/api/personas/<id>',
            'estadisticas': '/api/estadisticas'
        }
    }), 200

# ==================== INICIAR SERVIDOR ====================
if __name__ == '__main__':
    import os
    # Detectar si estamos en producción (Docker) o desarrollo
    is_production = os.getenv('FLASK_ENV') == 'production' or os.getenv('DOCKER_ENV') == 'true'
    
    print("=" * 60)
    print("🚕 API InfoTaxi iniciando...")
    print("=" * 60)
    print(f"📡 Servidor: http://0.0.0.0:5000")
    print(f"📚 Documentación Swagger: http://0.0.0.0:5000/apidocs/")
    print(f"🔍 Health Check: http://0.0.0.0:5000/api/health")
    print(f"🔧 Modo: {'Producción' if is_production else 'Desarrollo'}")
    print("=" * 60)
    
    # En producción usar host 0.0.0.0, en desarrollo localhost
    # debug solo en desarrollo
    app.run(
        debug=not is_production,
        host='0.0.0.0',
        port=int(os.getenv('PORT', 5000)),
        threaded=True
    )